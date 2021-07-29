#!/usr/bin/env python

import pandas as pd
import base64
from openpyxl import load_workbook
import streamlit as st
import os

'''
requires pandas==1.2.4
https://discuss.streamlit.io/t/change-background-color-based-on-value/2614/2
'''

def _max_width_():
    max_width_str = f"max-width: 1900px;"
    st.markdown(
        f"""
    <style>
    .reportview-container .main .block-container{{
        {max_width_str}
    }}
    </style>    
    """,
        unsafe_allow_html=True,
    )

_max_width_()

book = load_workbook(os.path.join("data","report.xlsx"))
st.title("Report Dashboard")
appnum=st.selectbox("Select App No.",book.sheetnames)

try:
    df = pd.read_excel(os.path.join("data","report.xlsx"))

    def highlight_validity(s):
        return ['background-color: green'] * len(s) if s.Validity > 90 else ['background-color: red'] * len(s)

    def color_validity(val):
        color = 'green' if val > 90 else 'red'
        return f'background-color: {color}'

    df.fillna(" ",inplace=True)
    # df.rename(columns={"Unnamed: 0":"Deployment Env"},inplace=True)
    df = df [['Status', 'Common Name', 'Valid From', 'Valid To',
       'Deployment Env', 'Device', 'Owner Email', 'App_Number',
       'License Status', 'Month', 'Year', 'Validity', 'app']]

    if len(df) > 0:
        # st.dataframe(df)
        st.dataframe(df.style.apply(highlight_validity,axis=1))
        st.dataframe(df.style.applymap(color_validity, subset=["Validity"]))
except Exception as err:
    print(err)
    st.error("No Data Found")


